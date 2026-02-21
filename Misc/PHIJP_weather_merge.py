import openpyxl
import pandas as pd
from openpyxl import load_workbook
import os
import time

start = time.time()
os.chdir('C:/Users/amitc_crl/OneDrive/Documents/GitHub/IIT-PhD/Misc')
wb = load_workbook(filename='Data List-Master_v2.xlsx')

ws = wb.active
steps = list(range(5,12073,12))
data = pd.DataFrame(columns=['Name',
                             '1_AT','2_AT','3_AT','4_AT','5_AT','6_AT','7_AT','8_AT','9_AT','10_AT','11_AT','12_AT',
                            '1_DP','2_DP','3_DP','4_DP','5_DP','6_DP','7_DP','8_DP','9_DP','10_DP','11_DP','12_DP',
                             '1_ST','2_ST','3_ST','4_ST','5_ST','6_ST','7_ST','8_ST','9_ST','10_ST','11_ST','12_ST',
                             '1_NR','2_NR','3_NR','4_NR','5_NR','6_NR','7_NR','8_NR','9_NR','10_NR','11_NR','12_NR',
                             '1_ER','2_ER','3_ER','4_ER','5_ER','6_ER','7_ER','8_ER','9_ER','10_ER','11_ER','12_ER',
                             '1_SR','2_SR','3_SR','4_SR','5_SR','6_SR','7_SR','8_SR','9_SR','10_SR','11_SR','12_SR',
                             '1_WR','2_WR','3_WR','4_WR','5_WR','6_WR','7_WR','8_WR','9_WR','10_WR','11_WR','12_WR',
                             '1_GR','2_GR','3_GR','4_GR','5_GR','6_GR','7_GR','8_GR','9_GR','10_GR','11_GR','12_GR',
                             'HW_1T','HW_2T','CW_1T',
                             'HW_1N','HW_2N','CW_1N',
                             'HW_1E','HW_2E','CW_1E',
                             'HW_1S','HW_2S','CW_1S',
                             'HW_1W','HW_2W','CW_1W',
                             'HW_1G','HW_2G','CW_1G','Swing_Temp'])

cols_1 = ['B','C','D','E','F','G','H','I','J','K','L','M']
cols_2 = ['N','O','P']
for step in steps:
    name = ws[str('A' + str(step))].value
    print(name)
    row = []
    row.append(name)

    for n in range(0,12,1):
        col = cols_1[n]
        row.append(ws[str(col)+str(step+1)].value)
    for n in range(0,12,1):
        col = cols_1[n]
        row.append(ws[str(col)+str(step+7)].value)
    for n in range(0,12,1):
        col = cols_1[n]
        row.append(ws[str(col)+str(step+8)].value)
    for n in range(0,12,1):
        col = cols_1[n]
        row.append(ws[str(col)+str(step+2)].value)
    for n in range(0,12,1):
        col = cols_1[n]
        row.append(ws[str(col)+str(step+3)].value)
    for n in range(0,12,1):
        col = cols_1[n]
        row.append(ws[str(col)+str(step+4)].value)
    for n in range(0,12,1):
        col = cols_1[n]
        row.append(ws[str(col)+str(step+5)].value)
    for n in range(0,12,1):
        col = cols_1[n]
        row.append(ws[str(col)+str(step+6)].value)
    
    for x in range(0,3,1):
        col = cols_2[x]
        row.append(ws[str(col)+str(step+1)].value)
    for x in range(0,3,1):
        col = cols_2[x]
        row.append(ws[str(col)+str(step+2)].value)
    for x in range(0,3,1):
        col = cols_2[x]
        row.append(ws[str(col)+str(step+3)].value)
    for x in range(0,3,1):
        col = cols_2[x]
        row.append(ws[str(col)+str(step+4)].value)
    for x in range(0,3,1):
        col = cols_2[x]
        row.append(ws[str(col)+str(step+5)].value)
    for x in range(0,3,1):
        col = cols_2[x]
        row.append(ws[str(col)+str(step+6)].value)
    row.append(ws[str('K')+str(step)].value)

    data.loc[len(data)] = row

data.to_csv('mergedData2.csv')

end = time.time()

print('Successful Completion')
print('Elapsed time = ' + str(end-start))